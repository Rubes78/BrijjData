import io
import statistics
from collections import OrderedDict

import openpyxl


class NoDataError(Exception):
    pass


def _clear_data_rows(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None


def build_department_workbook(template_bytes, items):
    depts = OrderedDict()
    for it in items:
        code = (it.get("category_code") or "").strip()
        if not code:
            continue
        desc = (it.get("category_code_description") or "").strip()
        if code not in depts or (desc and not depts[code]):
            depts[code] = desc

    if not depts:
        raise NoDataError("No items with a category_code were returned — nothing to sync.")

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    _clear_data_rows(ws)

    row_num = 2
    sort_order = 10
    for code in sorted(depts):
        desc = depts[code]
        name = desc if desc else code.title()
        ws.cell(row=row_num, column=1).value = name
        ws.cell(row=row_num, column=2).value = desc or name
        ws.cell(row=row_num, column=3).value = code
        ws.cell(row=row_num, column=4).value = sort_order
        ws.cell(row=row_num, column=5).value = None  # IconFileName — not sourced from rcsaero
        ws.cell(row=row_num, column=6).value = True  # TPM
        ws.cell(row=row_num, column=7).value = True  # POS
        ws.cell(row=row_num, column=8).value = True  # Lister
        ws.cell(row=row_num, column=9).value = True  # Donation
        row_num += 1
        sort_order += 10

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), len(depts)


def build_category_workbook(template_bytes, items):
    cats = OrderedDict()  # (dept_code, sub_code) -> (sub_desc, dept_desc)
    for it in items:
        dept_code = (it.get("category_code") or "").strip()
        sub_code = (it.get("subcategory_code") or "").strip()
        if not dept_code or not sub_code:
            continue
        sub_desc = (it.get("subcategory_code_description") or "").strip()
        dept_desc = (it.get("category_code_description") or "").strip()
        key = (dept_code, sub_code)
        if key not in cats:
            cats[key] = (sub_desc, dept_desc)

    if not cats:
        raise NoDataError(
            "No items had both a category_code and subcategory_code set — "
            "there is no subcategory data in this company's rcsaero items to sync."
        )

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    _clear_data_rows(ws)

    row_num = 2
    for (dept_code, sub_code), (sub_desc, dept_desc) in sorted(cats.items()):
        name = sub_desc if sub_desc else sub_code.title()
        product_department = dept_desc if dept_desc else dept_code.title()
        ws.cell(row=row_num, column=1).value = name
        ws.cell(row=row_num, column=2).value = sub_desc or name
        ws.cell(row=row_num, column=3).value = f"{dept_code}-{sub_code}"
        ws.cell(row=row_num, column=4).value = product_department
        ws.cell(row=row_num, column=5).value = None  # Category_Type — not sourced from rcsaero
        ws.cell(row=row_num, column=6).value = None  # Parent_Category
        ws.cell(row=row_num, column=7).value = "ZebraTag"
        ws.cell(row=row_num, column=8).value = None  # IconFileName — not sourced from rcsaero
        ws.cell(row=row_num, column=9).value = "Y"  # TPM
        ws.cell(row=row_num, column=10).value = "Y"  # POS
        ws.cell(row=row_num, column=11).value = "N"  # DONATION
        ws.cell(row=row_num, column=12).value = "N"  # LISTER
        row_num += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), len(cats)


def list_departments(items):
    """Distinct (code, description) departments present in a set of rcsaero items,
    sorted by code. Used to drive the pricing-department selection UI."""
    depts = OrderedDict()
    for it in items:
        code = (it.get("category_code") or "").strip()
        if not code:
            continue
        desc = (it.get("category_code_description") or "").strip()
        if code not in depts or (desc and not depts[code]):
            depts[code] = desc
    return [{"code": code, "description": depts[code] or code.title()} for code in sorted(depts)]


def list_categories(items):
    """Distinct (dept_code, sub_code) categories present in a set of rcsaero items
    (only items with both set), sorted. Used to drive the pricing-category selection UI."""
    cats = OrderedDict()  # composite_code -> (department_name, category_name)
    for it in items:
        dept_code = (it.get("category_code") or "").strip()
        sub_code = (it.get("subcategory_code") or "").strip()
        if not dept_code or not sub_code:
            continue
        dept_desc = (it.get("category_code_description") or "").strip()
        sub_desc = (it.get("subcategory_code_description") or "").strip()
        composite = f"{dept_code}-{sub_code}"
        if composite not in cats:
            cats[composite] = (dept_desc or dept_code.title(), sub_desc or sub_code.title())
    return [
        {"code": composite, "department": dept_name, "description": cat_name}
        for composite, (dept_name, cat_name) in sorted(cats.items())
    ]


PRICE_LIST_SHEET_NAME = "Price Lists by Department"
GBB_SHEET_NAME = "GBB by Department"
QUALITY_CONDITION_SHEET_NAME = "Quality and Condition by Depart"
CATEGORY_PRICE_LIST_SHEET_NAME = "Price Lists by Category and Sub"
CATEGORY_GBB_SHEET_NAME = "GBB by Category and SubCategory"
CATEGORY_QUALITY_CONDITION_SHEET_NAME = "Quality and Condition by Catego"
DEFAULT_STORE_GROUP = "Standard"

MODEL_LIST = "list"
MODEL_GBB = "gbb"
MODEL_QUALITY_CONDITION = "quality_condition"
VALID_MODELS = {MODEL_LIST, MODEL_GBB, MODEL_QUALITY_CONDITION}


def _collect_department_prices(items):
    """Returns {category_code: {"name": dept_name, "prices": [price, ...]}}
    for items with a price_1 set."""
    result = OrderedDict()
    for it in items:
        code = (it.get("category_code") or "").strip()
        if not code:
            continue
        price = it.get("price_1")
        if price is None:
            continue
        desc = (it.get("category_code_description") or "").strip()
        name = desc if desc else code.title()
        entry = result.setdefault(code, {"name": name, "prices": []})
        entry["prices"].append(float(price))
    return result


def _collect_category_prices(items):
    """Returns {f"{dept_code}-{sub_code}": {"department": dept_name, "name": cat_name,
    "prices": [...]}} for items with both category_code, subcategory_code, and price_1 set."""
    result = OrderedDict()
    for it in items:
        dept_code = (it.get("category_code") or "").strip()
        sub_code = (it.get("subcategory_code") or "").strip()
        if not dept_code or not sub_code:
            continue
        price = it.get("price_1")
        if price is None:
            continue
        dept_desc = (it.get("category_code_description") or "").strip()
        sub_desc = (it.get("subcategory_code_description") or "").strip()
        composite = f"{dept_code}-{sub_code}"
        entry = result.setdefault(composite, {
            "department": dept_desc or dept_code.title(),
            "name": sub_desc or sub_code.title(),
            "prices": [],
        })
        entry["prices"].append(float(price))
    return result


def _bucket_groups(prices, k):
    """Split a sorted list of prices into k equal-count buckets (quantile-style)
    and return the k price-lists in ascending order. Buckets can be empty if
    there are fewer observations than k."""
    prices = sorted(prices)
    n = len(prices)
    sizes = [n // k + (1 if i < n % k else 0) for i in range(k)]

    groups = []
    idx = 0
    for size in sizes:
        groups.append(prices[idx:idx + size])
        idx += size
    return groups


def _bucket_medians(prices, levels):
    """Split a sorted list of prices into `levels` equal-count buckets and
    return the median of each bucket, ascending. Empty buckets (fewer price
    observations than levels) come back as None."""
    return [round(statistics.median(g), 2) if g else None for g in _bucket_groups(prices, levels)]


def _list_pricing_prices(prices):
    """Every distinct price point, ascending — the List Pricing model has no tiers."""
    return sorted({round(p, 2) for p in prices})


def _gbb_tiers(prices, labels):
    """[(label, price), ...] — one row per non-empty equal-count bucket."""
    medians = _bucket_medians(prices, len(labels))
    return [(label, price) for label, price in zip(labels, medians) if price is not None]


def _quality_condition_tiers(prices, quality_labels, condition_labels):
    """[(quality_label, condition_label, price), ...]. Prices are bucketed into
    Quality tiers first, then each Quality bucket is further bucketed into
    Condition tiers — mirroring the ascending Quality-then-Condition price
    progression already used in the hand-built Quality and Condition by Category sheet."""
    rows = []
    for quality_label, quality_group in zip(quality_labels, _bucket_groups(prices, len(quality_labels))):
        if not quality_group:
            continue
        for condition_label, price in zip(condition_labels, _bucket_medians(quality_group, len(condition_labels))):
            if price is not None:
                rows.append((quality_label, condition_label, price))
    return rows


def _write_rows(ws, next_row, prefix, rows):
    """Writes `prefix` (a fixed list of leading column values, e.g. [Store Group, Department])
    followed by each tuple in `rows` (the model-specific trailing columns) as one row per tuple.
    Returns the updated next_row and how many rows were written."""
    written = 0
    for trailing in rows:
        for i, value in enumerate(prefix, start=1):
            ws.cell(row=next_row, column=i).value = value
        for j, value in enumerate(trailing, start=len(prefix) + 1):
            ws.cell(row=next_row, column=j).value = value
        next_row += 1
        written += 1
    return next_row, written


def _write_priced_entity(ws_by_model, next_row, prefix, model, prices, gbb_labels, condition_labels):
    """Dispatches one department/category's prices to the right sheet based on `model`,
    writing `prefix` + the model's trailing columns for each resulting row."""
    if model == MODEL_LIST:
        rows = [(price,) for price in _list_pricing_prices(prices)]
    elif model == MODEL_GBB:
        rows = list(_gbb_tiers(prices, gbb_labels))
    else:  # MODEL_QUALITY_CONDITION
        rows = list(_quality_condition_tiers(prices, gbb_labels, condition_labels))

    ws = ws_by_model[model]
    r, written = _write_rows(ws, next_row[model], prefix, rows)
    next_row[model] = r
    return written


def build_combined_pricing_workbook(
    template_bytes, items, department_models, category_models, gbb_labels, condition_labels
):
    """department_models / category_models: {code: "list" | "gbb" | "quality_condition"}.
    department code is the rcsaero category_code; category code is "{category_code}-{subcategory_code}".
    Each department/category gets exactly one pricing model — its rows go into exactly one
    of the six sheets below (three department-level, three category-level). Anything absent
    from its models dict (or mapped to something invalid) is skipped entirely, on every sheet."""
    dept_data = _collect_department_prices(items)
    cat_data = _collect_category_prices(items)

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    all_sheet_names = (
        PRICE_LIST_SHEET_NAME, GBB_SHEET_NAME, QUALITY_CONDITION_SHEET_NAME,
        CATEGORY_PRICE_LIST_SHEET_NAME, CATEGORY_GBB_SHEET_NAME, CATEGORY_QUALITY_CONDITION_SHEET_NAME,
    )
    sheets = {}
    for name in all_sheet_names:
        if name not in wb.sheetnames:
            raise NoDataError(f'Pricing template is missing the "{name}" sheet.')
        sheets[name] = wb[name]
        _clear_data_rows(sheets[name])

    dept_ws_by_model = {
        MODEL_LIST: sheets[PRICE_LIST_SHEET_NAME],
        MODEL_GBB: sheets[GBB_SHEET_NAME],
        MODEL_QUALITY_CONDITION: sheets[QUALITY_CONDITION_SHEET_NAME],
    }
    cat_ws_by_model = {
        MODEL_LIST: sheets[CATEGORY_PRICE_LIST_SHEET_NAME],
        MODEL_GBB: sheets[CATEGORY_GBB_SHEET_NAME],
        MODEL_QUALITY_CONDITION: sheets[CATEGORY_QUALITY_CONDITION_SHEET_NAME],
    }
    dept_next_row = {MODEL_LIST: 2, MODEL_GBB: 2, MODEL_QUALITY_CONDITION: 2}
    cat_next_row = {MODEL_LIST: 2, MODEL_GBB: 2, MODEL_QUALITY_CONDITION: 2}
    row_count = 0

    for code in sorted(dept_data):
        model = department_models.get(code)
        if model not in VALID_MODELS:
            continue
        entry = dept_data[code]
        prefix = [DEFAULT_STORE_GROUP, entry["name"]]
        row_count += _write_priced_entity(
            dept_ws_by_model, dept_next_row, prefix, model, entry["prices"], gbb_labels, condition_labels
        )

    for code in sorted(cat_data):
        model = category_models.get(code)
        if model not in VALID_MODELS:
            continue
        entry = cat_data[code]
        prefix = [DEFAULT_STORE_GROUP, entry["department"], entry["name"], None]  # Sub-Category left blank
        row_count += _write_priced_entity(
            cat_ws_by_model, cat_next_row, prefix, model, entry["prices"], gbb_labels, condition_labels
        )

    if row_count == 0:
        raise NoDataError(
            "No departments or categories were assigned a pricing model "
            "(or none had enough price data) — nothing to sync."
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), row_count
